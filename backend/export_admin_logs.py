"""
Admin Logs를 Excel로 내보내는 스크립트
- Supabase admin_logs 테이블에서 전체 데이터 조회
- conversation_history (잘리지 않은 전체), user_question, router_output, function_result, final_answer를 엑셀로 저장
"""

import os
import json
from datetime import datetime
from dotenv import load_dotenv
from supabase import create_client

# openpyxl 설치 필요: pip install openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   pip install openpyxl 실행 후 다시 시도하세요.")
    exit(1)

# 환경 변수 로드
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ SUPABASE_URL 또는 SUPABASE_KEY가 설정되지 않았습니다.")
    exit(1)


def fetch_all_logs():
    """Supabase에서 모든 admin_logs 조회 (페이지네이션 처리)"""
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    all_logs = []
    page_size = 1000
    offset = 0
    
    print("📥 Supabase에서 로그 조회 중...")
    
    while True:
        result = client.table('admin_logs') \
            .select('id, timestamp, conversation_history, user_question, router_output, function_result, final_answer') \
            .order('timestamp', desc=True) \
            .range(offset, offset + page_size - 1) \
            .execute()
        
        if not result.data:
            break
        
        all_logs.extend(result.data)
        print(f"   {len(all_logs)}개 로드됨...")
        
        if len(result.data) < page_size:
            break
        
        offset += page_size
    
    print(f"✅ 총 {len(all_logs)}개 로그 조회 완료")
    return all_logs


def format_conversation_history(history):
    """conversation_history를 읽기 좋은 문자열로 변환"""
    if not history:
        return ""
    
    if isinstance(history, str):
        try:
            history = json.loads(history)
        except:
            return history
    
    if isinstance(history, list):
        # 리스트 형태: ["user: 질문1", "assistant: 답변1", ...]
        return "\n".join(history)
    
    return str(history)


def format_json_field(data):
    """JSON 필드를 읽기 좋은 문자열로 변환"""
    if not data:
        return ""
    
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except:
            return data
    
    try:
        return json.dumps(data, ensure_ascii=False, indent=2)
    except:
        return str(data)


def export_to_excel(logs, output_path):
    """로그를 Excel 파일로 내보내기"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Logs"
    
    # 헤더 설정
    headers = [
        "ID",
        "Timestamp",
        "1. Conversation History (전체)",
        "2. User Question",
        "3. Router Output (JSON)",
        "4. Function Result (청크)",
        "5. Final Answer"
    ]
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 입력
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    print("📝 Excel 파일 생성 중...")
    
    for row_idx, log in enumerate(logs, 2):
        # ID
        ws.cell(row=row_idx, column=1, value=log.get('id', ''))
        
        # Timestamp
        ws.cell(row=row_idx, column=2, value=log.get('timestamp', ''))
        
        # 1. Conversation History (전체)
        history = format_conversation_history(log.get('conversation_history'))
        cell = ws.cell(row=row_idx, column=3, value=history)
        cell.alignment = cell_alignment
        
        # 2. User Question
        cell = ws.cell(row=row_idx, column=4, value=log.get('user_question', ''))
        cell.alignment = cell_alignment
        
        # 3. Router Output
        router_output = format_json_field(log.get('router_output'))
        cell = ws.cell(row=row_idx, column=5, value=router_output)
        cell.alignment = cell_alignment
        
        # 4. Function Result
        function_result = format_json_field(log.get('function_result'))
        cell = ws.cell(row=row_idx, column=6, value=function_result)
        cell.alignment = cell_alignment
        
        # 5. Final Answer
        cell = ws.cell(row=row_idx, column=7, value=log.get('final_answer', ''))
        cell.alignment = cell_alignment
        
        if row_idx % 100 == 0:
            print(f"   {row_idx - 1}개 처리됨...")
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 10  # ID
    ws.column_dimensions['B'].width = 22  # Timestamp
    ws.column_dimensions['C'].width = 60  # Conversation History
    ws.column_dimensions['D'].width = 50  # User Question
    ws.column_dimensions['E'].width = 50  # Router Output
    ws.column_dimensions['F'].width = 80  # Function Result
    ws.column_dimensions['G'].width = 80  # Final Answer
    
    # 첫 행 고정
    ws.freeze_panes = 'A2'
    
    # 저장
    wb.save(output_path)
    print(f"✅ Excel 파일 저장 완료: {output_path}")


def main():
    print("=" * 60)
    print("Admin Logs Excel 내보내기")
    print("=" * 60)
    
    # 로그 조회
    logs = fetch_all_logs()
    
    if not logs:
        print("⚠️ 조회된 로그가 없습니다.")
        return
    
    # 출력 파일명 (현재 시간 포함)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"admin_logs_export_{timestamp}.xlsx"
    
    # Excel 내보내기
    export_to_excel(logs, output_path)
    
    print("\n" + "=" * 60)
    print(f"📊 총 {len(logs)}개 로그가 '{output_path}'에 저장되었습니다.")
    print("=" * 60)


if __name__ == "__main__":
    main()
