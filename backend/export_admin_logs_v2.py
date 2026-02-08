"""
Admin Logs를 Excel로 내보내는 스크립트 (v2)
- admin_logs 테이블에서 기본 데이터 조회
- session_chat_messages 테이블에서 전체 대화 히스토리 조회 (잘리지 않은 전체)
- user_id + timestamp 기준으로 매칭하여 완전한 히스토리 복원
"""

import os
import json
from datetime import datetime, timedelta
from dotenv import load_dotenv
from supabase import create_client

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   pip install openpyxl 실행 후 다시 시도하세요.")
    exit(1)

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ SUPABASE_URL 또는 SUPABASE_KEY가 설정되지 않았습니다.")
    exit(1)


def fetch_all_admin_logs():
    """Supabase에서 모든 admin_logs 조회"""
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    all_logs = []
    page_size = 1000
    offset = 0
    
    print("📥 admin_logs 테이블에서 로그 조회 중...")
    
    while True:
        result = client.table('admin_logs') \
            .select('id, user_id, timestamp, conversation_history, user_question, router_output, function_result, final_answer') \
            .order('timestamp', desc=True) \
            .range(offset, offset + page_size - 1) \
            .execute()
        
        if not result.data:
            break
        
        all_logs.extend(result.data)
        print(f"   {len(all_logs)}개 로드됨...")
        
        if len(result.data) < page_size:
            break
        
        offset += page_size
    
    print(f"✅ 총 {len(all_logs)}개 admin_logs 조회 완료")
    return all_logs


def fetch_all_session_messages():
    """session_chat_messages 테이블에서 전체 대화 조회"""
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    all_messages = []
    page_size = 1000
    offset = 0
    
    print("📥 session_chat_messages 테이블에서 대화 조회 중...")
    
    while True:
        result = client.table('session_chat_messages') \
            .select('user_id, user_session, role, content, created_at') \
            .order('created_at', desc=False) \
            .range(offset, offset + page_size - 1) \
            .execute()
        
        if not result.data:
            break
        
        all_messages.extend(result.data)
        print(f"   {len(all_messages)}개 로드됨...")
        
        if len(result.data) < page_size:
            break
        
        offset += page_size
    
    print(f"✅ 총 {len(all_messages)}개 session_chat_messages 조회 완료")
    return all_messages


def build_session_history_map(messages):
    """
    session_chat_messages를 user_id + user_session 기준으로 그룹화
    각 세션의 대화 히스토리를 시간순으로 정렬
    """
    print("🔧 세션별 대화 히스토리 구축 중...")
    
    # user_id -> user_session -> messages list
    history_map = {}
    
    for msg in messages:
        user_id = msg.get('user_id')
        session_id = msg.get('user_session')
        
        if not user_id or not session_id:
            continue
        
        if user_id not in history_map:
            history_map[user_id] = {}
        
        if session_id not in history_map[user_id]:
            history_map[user_id][session_id] = []
        
        history_map[user_id][session_id].append({
            'role': msg.get('role', ''),
            'content': msg.get('content', ''),
            'created_at': msg.get('created_at', '')
        })
    
    # 각 세션 내 메시지를 시간순 정렬
    for user_id in history_map:
        for session_id in history_map[user_id]:
            history_map[user_id][session_id].sort(key=lambda x: x['created_at'])
    
    print(f"✅ {len(history_map)}명의 사용자, 총 {sum(len(sessions) for sessions in history_map.values())}개 세션 구축 완료")
    return history_map


def find_matching_history(log, history_map):
    """
    admin_log에 해당하는 전체 대화 히스토리 찾기
    - user_id로 사용자 매칭
    - timestamp 기준으로 해당 시점까지의 대화 찾기
    """
    user_id = log.get('user_id')
    log_timestamp = log.get('timestamp')
    user_question = log.get('user_question', '')
    
    if not user_id or user_id not in history_map:
        return None
    
    user_sessions = history_map[user_id]
    
    # 모든 세션에서 해당 질문이 포함된 세션 찾기
    best_match = None
    best_session_id = None
    
    for session_id, messages in user_sessions.items():
        # 해당 세션에서 user_question과 일치하는 메시지 찾기
        for i, msg in enumerate(messages):
            if msg['role'] == 'user' and user_question in msg['content']:
                # 해당 메시지까지의 히스토리 반환
                history_up_to_now = messages[:i+1]
                
                # 더 긴 히스토리를 우선
                if best_match is None or len(history_up_to_now) > len(best_match):
                    best_match = history_up_to_now
                    best_session_id = session_id
    
    return best_match


def format_full_history(messages):
    """전체 대화 히스토리를 읽기 좋은 문자열로 변환"""
    if not messages:
        return ""
    
    lines = []
    for msg in messages:
        role = "User" if msg['role'] == 'user' else "Bot"
        content = msg['content'] or ''
        lines.append(f"{role}: {content}")
    
    return "\n".join(lines)


def format_json_field(data):
    """JSON 필드를 읽기 좋은 문자열로 변환"""
    if not data:
        return ""
    
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except:
            return data
    
    try:
        return json.dumps(data, ensure_ascii=False, indent=2)
    except:
        return str(data)


def export_to_excel(logs, history_map, output_path):
    """로그를 Excel 파일로 내보내기"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Logs (Full History)"
    
    headers = [
        "ID",
        "Timestamp",
        "User ID",
        "1. Conversation History (전체 - session_chat_messages)",
        "2. User Question",
        "3. Router Output (JSON)",
        "4. Function Result (청크)",
        "5. Final Answer"
    ]
    
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    print("📝 Excel 파일 생성 중...")
    
    matched_count = 0
    unmatched_count = 0
    
    for row_idx, log in enumerate(logs, 2):
        # ID
        ws.cell(row=row_idx, column=1, value=log.get('id', ''))
        
        # Timestamp
        ws.cell(row=row_idx, column=2, value=log.get('timestamp', ''))
        
        # User ID
        ws.cell(row=row_idx, column=3, value=log.get('user_id', ''))
        
        # 1. Conversation History (전체) - session_chat_messages에서 매칭
        full_history = find_matching_history(log, history_map)
        if full_history:
            history_text = format_full_history(full_history)
            matched_count += 1
        else:
            # 매칭 실패 시 기존 conversation_history 사용 (잘린 버전)
            old_history = log.get('conversation_history', [])
            if isinstance(old_history, list):
                history_text = "\n".join(old_history) if old_history else "(히스토리 없음)"
            else:
                history_text = str(old_history) if old_history else "(히스토리 없음)"
            unmatched_count += 1
        
        cell = ws.cell(row=row_idx, column=4, value=history_text)
        cell.alignment = cell_alignment
        
        # 2. User Question
        cell = ws.cell(row=row_idx, column=5, value=log.get('user_question', ''))
        cell.alignment = cell_alignment
        
        # 3. Router Output
        router_output = format_json_field(log.get('router_output'))
        cell = ws.cell(row=row_idx, column=6, value=router_output)
        cell.alignment = cell_alignment
        
        # 4. Function Result
        function_result = format_json_field(log.get('function_result'))
        cell = ws.cell(row=row_idx, column=7, value=function_result)
        cell.alignment = cell_alignment
        
        # 5. Final Answer
        cell = ws.cell(row=row_idx, column=8, value=log.get('final_answer', ''))
        cell.alignment = cell_alignment
        
        if row_idx % 100 == 0:
            print(f"   {row_idx - 1}개 처리됨... (매칭: {matched_count}, 미매칭: {unmatched_count})")
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = 10   # ID
    ws.column_dimensions['B'].width = 22   # Timestamp
    ws.column_dimensions['C'].width = 38   # User ID
    ws.column_dimensions['D'].width = 80   # Conversation History (전체)
    ws.column_dimensions['E'].width = 50   # User Question
    ws.column_dimensions['F'].width = 50   # Router Output
    ws.column_dimensions['G'].width = 80   # Function Result
    ws.column_dimensions['H'].width = 80   # Final Answer
    
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"✅ Excel 파일 저장 완료: {output_path}")
    print(f"   - 전체 히스토리 매칭 성공: {matched_count}개")
    print(f"   - 매칭 실패 (기존 히스토리 사용): {unmatched_count}개")


def main():
    print("=" * 60)
    print("Admin Logs Excel 내보내기 (v2 - 전체 히스토리)")
    print("=" * 60)
    
    # 1. admin_logs 조회
    logs = fetch_all_admin_logs()
    if not logs:
        print("⚠️ 조회된 로그가 없습니다.")
        return
    
    # 2. session_chat_messages 조회
    messages = fetch_all_session_messages()
    
    # 3. 세션별 히스토리 맵 구축
    history_map = build_session_history_map(messages)
    
    # 4. 출력 파일명
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"admin_logs_full_history_{timestamp}.xlsx"
    
    # 5. Excel 내보내기
    export_to_excel(logs, history_map, output_path)
    
    print("\n" + "=" * 60)
    print(f"📊 총 {len(logs)}개 로그가 '{output_path}'에 저장되었습니다.")
    print("=" * 60)


if __name__ == "__main__":
    main()
