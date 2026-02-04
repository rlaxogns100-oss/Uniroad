"""
Excel 출력 유틸리티
- 테스트 결과를 Excel 파일로 내보내기
- Pipeline용 상세 출력 지원
"""

import os
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def export_results_to_excel(
    results: List[Dict[str, Any]],
    output_dir: str,
    test_name: str = None
) -> str:
    """
    테스트 결과를 Excel 파일로 내보내기
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"test_result_{test_name}_{timestamp}.xlsx" if test_name else f"test_result_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Pipeline인지 확인
    is_pipeline = any(r.get("agent_type") == "pipeline" for r in results)
    
    if is_pipeline:
        return _export_pipeline_results(results, filepath)
    else:
        return _export_standard_results(results, filepath)


def _export_pipeline_results(results: List[Dict[str, Any]], filepath: str) -> str:
    """Pipeline 전용 상세 Excel 출력 - 간단 버전"""
    wb = Workbook()
    
    # 메인 결과 시트
    ws_main = wb.active
    ws_main.title = "Pipeline Results"
    
    # 헤더 정의 (세분화)
    headers = [
        "Run_ID",
        "Success",
        "Total_Time_ms",
        "Input_Full",
        "Orch_User_Intent",
        "Orch_Extracted_Scores",
        "Orch_Execution_Plan",
        "Orch_Answer_Structure",
        "Orch_Time_ms",
        "Sub_Results",
        "Sub_Time_ms",
        "Final_Output",
        "Final_Time_ms",
        "Error",
        "Timestamp"
    ]
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 헤더 추가
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 추가
    for row_idx, r in enumerate(results, 2):
        # Input 전체 (대화 기록 + 질문)
        input_history_raw = r.get('input_history', [])
        # input_history가 문자열이면 파싱
        if isinstance(input_history_raw, str):
            try:
                import json
                input_history = json.loads(input_history_raw)
            except:
                input_history = []
        else:
            input_history = input_history_raw
        
        input_message = r.get('input_message', '')
        full_input = ""
        if input_history:
            full_input += "=== 대화 기록 ===\n"
            for msg in input_history:
                role = msg.get('role', 'unknown')
                content = msg.get('content', '')
                full_input += f"[{role}]: {content}\n\n"
        full_input += f"=== 현재 질문 ===\n{input_message}"
        
        row_data = [
            r.get("run_id", 0),
            "Y" if r.get("success", False) else "N",
            r.get("processing_time_ms", 0),
            full_input,
            r.get('orchestration_user_intent', ''),
            r.get('orchestration_extracted_scores', ''),
            r.get('orchestration_execution_plan', ''),
            r.get('orchestration_answer_structure', ''),
            r.get('orchestration_time_ms', 0),
            r.get('sub_agents_results', ''),
            r.get('sub_agents_time_ms', 0),
            r.get('final_output', ''),
            r.get('final_time_ms', 0),
            r.get("error", ""),
            r.get("timestamp", "")
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Success 컬럼 색상
            if col_idx == 2:
                cell.fill = success_fill if value == "Y" else fail_fill
    
    # 컬럼 너비
    column_widths = {
        'A': 10,  # Run_ID
        'B': 10,  # Success
        'C': 15,  # Total_Time_ms
        'D': 80,  # Input_Full
        'E': 50,  # Orch_User_Intent
        'F': 50,  # Orch_Extracted_Scores
        'G': 80,  # Orch_Execution_Plan
        'H': 80,  # Orch_Answer_Structure
        'I': 12,  # Orch_Time_ms
        'J': 100, # Sub_Results
        'K': 12,  # Sub_Time_ms
        'L': 100, # Final_Output
        'M': 12,  # Final_Time_ms
        'N': 50,  # Error
        'O': 22   # Timestamp
    }
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # 통계 시트
    ws_stats = wb.create_sheet(title="Statistics")
    _add_statistics_sheet(ws_stats, results)
    
    # 저장
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath


def _export_orchestration_results(results: List[Dict[str, Any]], filepath: str) -> str:
    """Orchestration 전용 상세 Excel 출력"""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Orchestration Results"
    
    # 헤더 정의
    headers = [
        "Run_ID",
        "Success",
        "Processing_Time_ms",
        "Input_Summary",
        "User_Intent",
        "Extracted_Scores",
        "Execution_Plan",
        "Answer_Structure",
        "Error",
        "Timestamp",
        "Token_Usage"
    ]
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 헤더 추가
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 추가
    for row_idx, r in enumerate(results, 2):
        import json
        
        # output에서 orchestration 세부 정보 추출
        output = r.get("output", {})
        if isinstance(output, str):
            try:
                output = json.loads(output)
            except:
                output = {}
        
        row_data = [
            r.get("run_id", 0),
            "Y" if r.get("success", False) else "N",
            r.get("processing_time_ms", 0),
            r.get("input_summary", ""),
            output.get("user_intent", ""),
            json.dumps(output.get("extracted_scores", {}), ensure_ascii=False, indent=2) if output.get("extracted_scores") else "",
            json.dumps(output.get("execution_plan", []), ensure_ascii=False, indent=2),
            json.dumps(output.get("answer_structure", []), ensure_ascii=False, indent=2),
            r.get("error", ""),
            r.get("timestamp", ""),
            r.get("token_usage", 0)
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Success 컬럼 색상
            if col_idx == 2:
                cell.fill = success_fill if value == "Y" else fail_fill
    
    # 컬럼 너비
    column_widths = {
        'A': 10,  # Run_ID
        'B': 10,  # Success
        'C': 18,  # Processing_Time_ms
        'D': 50,  # Input_Summary
        'E': 50,  # User_Intent
        'F': 50,  # Extracted_Scores
        'G': 80,  # Execution_Plan
        'H': 80,  # Answer_Structure
        'I': 50,  # Error
        'J': 22,  # Timestamp
        'K': 15   # Token_Usage
    }
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # 통계 시트
    ws_stats = wb.create_sheet(title="Statistics")
    _add_statistics_sheet(ws_stats, results)
    
    # 저장
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath


def _export_sub_agents_results(results: List[Dict[str, Any]], filepath: str) -> str:
    """Sub Agents 전용 상세 Excel 출력 (각 에이전트별 행 분리)"""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Sub Agents Results"
    
    # 헤더 정의
    headers = [
        "Run_ID",
        "Agent_Name",
        "Step",
        "Query",
        "Retrieved_Data",
        "Status",
        "Processing_Time_ms",
        "Error",
        "Timestamp"
    ]
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 헤더 추가
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 추가 (각 에이전트별 행 분리)
    row_idx = 2
    for r in results:
        import json
        
        run_id = r.get("run_id", 0)
        processing_time_ms = r.get("processing_time_ms", 0)
        timestamp = r.get("timestamp", "")
        main_error = r.get("error", "")
        
        # output에서 sub agents 세부 정보 추출
        output = r.get("output", {})
        if isinstance(output, str):
            try:
                output = json.loads(output)
            except:
                output = {}
        
        # results 필드에서 실제 조회 데이터 추출
        sub_results = output.get("results", {})
        
        if not sub_results or not isinstance(sub_results, dict):
            # 결과가 없으면 에러 행 추가
            row_data = [
                run_id,
                "N/A",
                "",
                "",
                "",
                "error",
                processing_time_ms,
                main_error or "No results",
                timestamp
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_idx += 1
            continue
        
        # 각 스텝별로 행 생성
        for step_key, step_data in sub_results.items():
            if not isinstance(step_data, dict):
                continue
                
            agent_name = step_data.get("agent", step_key)
            query = step_data.get("query", "")
            result_text = step_data.get("result", "")
            status = step_data.get("status", "unknown")
            
            row_data = [
                run_id,
                agent_name,
                step_key,
                query,
                result_text,
                status,
                processing_time_ms,
                main_error,
                timestamp
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Status 컬럼 색상
                if col_idx == 6:
                    if value == "success":
                        cell.fill = success_fill
                    elif value == "error":
                        cell.fill = fail_fill
            
            row_idx += 1
    
    # 컬럼 너비
    column_widths = {
        'A': 10,  # Run_ID
        'B': 30,  # Agent_Name
        'C': 15,  # Step
        'D': 80,  # Query
        'E': 120, # Retrieved_Data (가장 중요!)
        'F': 12,  # Status
        'G': 18,  # Processing_Time_ms
        'H': 50,  # Error
        'I': 22   # Timestamp
    }
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # 통계 시트
    ws_stats = wb.create_sheet(title="Statistics")
    _add_statistics_sheet(ws_stats, results)
    
    # 저장
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath


def _export_final_agent_results(results: List[Dict[str, Any]], filepath: str) -> str:
    """Final Agent 전용 상세 Excel 출력"""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Final Agent Results"
    
    # 헤더 정의
    headers = [
        "Run_ID",
        "Success",
        "Processing_Time_ms",
        "Input_User_Question",
        "Input_History",
        "Answer_Structure",
        "Sub_Agent_Results",
        "Final_Output",
        "Error",
        "Timestamp",
        "Token_Usage"
    ]
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 헤더 추가
    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 추가
    for row_idx, r in enumerate(results, 2):
        import json
        
        # output에서 final agent 세부 정보 추출
        output = r.get("output", {})
        if isinstance(output, str):
            try:
                output = json.loads(output)
            except:
                pass
        
        # metadata에서 입력 데이터 추출
        metadata = r.get("metadata", {})
        user_question = metadata.get("user_question", "")
        history = metadata.get("history", [])
        answer_structure = metadata.get("answer_structure", [])
        sub_agent_results = metadata.get("sub_agent_results", {})
        
        # Final output 전체 (truncate 없음!)
        if isinstance(output, dict):
            final_output = output.get("final_answer", str(output))
        else:
            final_output = str(output)
        
        # JSON 포맷팅 (truncate 없음!)
        history_str = json.dumps(history, ensure_ascii=False, indent=2) if history else ""
        answer_structure_str = json.dumps(answer_structure, ensure_ascii=False, indent=2) if answer_structure else ""
        sub_agent_results_str = json.dumps(sub_agent_results, ensure_ascii=False, indent=2) if sub_agent_results else ""
        
        row_data = [
            r.get("run_id", 0),
            "Y" if r.get("success", False) else "N",
            r.get("processing_time_ms", 0),
            user_question,
            history_str,
            answer_structure_str,
            sub_agent_results_str,
            final_output,  # Full output, NO TRUNCATE!
            r.get("error", ""),
            r.get("timestamp", ""),
            r.get("token_usage", 0)
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Success 컬럼 색상
            if col_idx == 2:
                cell.fill = success_fill if value == "Y" else fail_fill
    
    # 컬럼 너비
    column_widths = {
        'A': 10,  # Run_ID
        'B': 10,  # Success
        'C': 18,  # Processing_Time_ms
        'D': 80,  # Input_User_Question
        'E': 80,  # Input_History
        'F': 80,  # Answer_Structure
        'G': 100, # Sub_Agent_Results
        'H': 120, # Final_Output (가장 중요! 전체 출력)
        'I': 50,  # Error
        'J': 22,  # Timestamp
        'K': 15   # Token_Usage
    }
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # 통계 시트
    ws_stats = wb.create_sheet(title="Statistics")
    _add_statistics_sheet(ws_stats, results)
    
    # 저장
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath


def _export_standard_results(results: List[Dict[str, Any]], filepath: str) -> str:
    """표준 Excel 출력 (Orchestration, Sub Agents, Final은 세분화)"""
    # Agent 타입 확인
    if results:
        agent_type = results[0].get("agent_type")
        if agent_type == "orchestration":
            return _export_orchestration_results(results, filepath)
        elif agent_type == "sub_agents":
            return _export_sub_agents_results(results, filepath)
        elif agent_type == "final":
            return _export_final_agent_results(results, filepath)
    
    # 일반 출력
    rows = []
    for result in results:
        row = {
            "Run_ID": result.get("run_id", 0),
            "Agent_Type": result.get("agent_type", "unknown"),
            "Model": result.get("model", "unknown"),
            "Step": result.get("step", ""),
            "Input_Summary": str(result.get("input_summary", ""))[:500],
            "Output": str(result.get("output", ""))[:2000],
            "Processing_Time_ms": result.get("processing_time_ms", 0),
            "Success": "Y" if result.get("success", False) else "N",
            "Timestamp": result.get("timestamp", ""),
            "Token_Usage": result.get("token_usage", 0),
            "Error": result.get("error", "")
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c_idx == 8:
                    if value == "Y":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    column_widths = {'A': 8, 'B': 18, 'C': 22, 'D': 8, 'E': 50, 'F': 80, 'G': 18, 'H': 10, 'I': 22, 'J': 12, 'K': 40}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws_stats = wb.create_sheet(title="Statistics")
    _add_statistics_sheet(ws_stats, results)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return filepath


def _add_statistics_sheet(ws, results: List[Dict[str, Any]]):
    """통계 시트 추가"""
    total_runs = len(results)
    success_count = sum(1 for r in results if r.get("success", False))
    times = [r.get("processing_time_ms", 0) for r in results]
    avg_time = sum(times) / len(times) if times else 0
    max_time = max(times) if times else 0
    min_time = min(times) if times else 0
    
    stats = [
        ("총 실행 횟수", total_runs),
        ("성공 횟수", success_count),
        ("실패 횟수", total_runs - success_count),
        ("성공률 (%)", round(success_count / total_runs * 100, 1) if total_runs > 0 else 0),
        ("", ""),
        ("평균 처리 시간 (ms)", round(avg_time, 2)),
        ("최대 처리 시간 (ms)", round(max_time, 2)),
        ("최소 처리 시간 (ms)", round(min_time, 2)),
    ]
    
    header_font = Font(bold=True)
    
    for row_idx, (label, value) in enumerate(stats, 1):
        ws.cell(row=row_idx, column=1, value=label).font = header_font
        ws.cell(row=row_idx, column=2, value=value)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
